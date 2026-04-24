# encoding: utf-8
"""
@version: 0-0
@author: huangpeng
@contact: huangpeng@aicsemi.com 
@site:
@python: 3.7.4
@software: PyCharm
@file: excel.py
@time: 2019/8/7 13:35
"""

import os
import sys
from loglib import *
from pybasic.hpfile import *
from pybasic.hpdecor import *
from openpyxl import load_workbook, workbook
from openpyxl.styles import PatternFill


class HpExcel(HpFile):
    def __init__(self, excel_file_path):
        super(HpExcel, self).__init__(excel_file_path)
        self.red_fill = PatternFill("solid", fgColor="FF0000")
        try:
            self.work_book = load_workbook(self.path)
        except Exception as e:
            pass

    @property
    def sheet_names(self):
        return self.work_book.sheetnames

    def sheet_maxrow(self, sheet_name):
        return self.work_book[sheet_name].max_row

    def sheet_maxcol(self, sheet_name):
        return self.work_book[sheet_name].max_column

    @decor_catch_exception
    def read_sheet_list(self, sheet_name):
        # return: content as a list
        content_list = []
        sheet = self.work_book[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column
        for row in range(1, max_row+1):
            row_list = []
            for col in range(1, max_col+1):
                row_list.append(sheet.cell(row, col).value)
            content_list.append(row_list)
        return content_list

    @decor_check_action
    def write_cell(self, sheet_name, cell_row, cell_col, cell_value):
        sheet = self.work_book[sheet_name]
        sheet.cell(cell_row, cell_col).value = cell_value
        self.work_book.save(self.path)

    @decor_catch_exception
    def read_cell(self, sheet_name: str, cell_row, cell_col) -> str:
        return self.work_book[sheet_name].cell(cell_row, cell_col).value

    @decor_check_action
    def set_cell_color(self, sheet_name, cell_row, cell_col, p_fill):
        self.work_book[sheet_name].cell(cell_row, cell_col).fill = p_fill
        self.work_book.save(self.path)


if __name__ == "__main__":
    excel_file = r"E:\scripts\dijiang\table\xxx_aic8000_bt_interface_pinlist_v055.xlsx"
    EXCEL1 = HpExcel(excel_file)
    sheetNames = EXCEL1.sheet_names
    sheet_name = sheetNames[0]
    content = EXCEL1.read_sheet_list(sheet_name)
    print(content)
    # content = EXCEL1.read()
