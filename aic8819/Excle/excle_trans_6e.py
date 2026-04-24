import pandas as pd
from openpyxl import load_workbook
from icbasic.toolkit.util import *
# 读取表1
df = pd.read_excel("apc22_6e_data_0928_all_swpV2.xlsx", sheet_name="Sheet1")

# 打开表2
wb = load_workbook(r"D:\Aic8800\8822\AIC8822_APC_6E_20250924.xlsx")
it_name = r'wf_dac0/1_vl_it_bit'
ib_name = r'wf_dac0/1_vl_ib_bit'

for _, row in df.iterrows():
    ch = int(row["ch_case"])
    freq = get_freq_by_ch_6e(ch)
    sheet_name = 'CH'+str(ch)+'_'+str(freq)+'M'
    gear = str(row["analog_gear"]).upper()   # 列名 (A, B, C ... 或数字转16进制)
    it_bit = row['wf_dac0_vl_it_bit']
    ib_bit = row['wf_dac0_vl_ib_bit']

    if sheet_name not in wb.sheetnames:
        print(f"⚠️ sheet {sheet_name} 不存在，跳过")
        continue

    ws = wb[sheet_name]

    # 找到行号（APC_BITS = wf_dtmx_dac_vl_it_bit / wf_dtmx_dac_vl_ib_bit）
    row_it = None
    row_ib = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val == it_name:
            row_it = r
        elif val == ib_name:
            row_ib = r

    # 找到列号（根据 analog_gear）
    col_idx = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value).upper() == gear:
            col_idx = c
            break

    if row_it and col_idx:
        ws.cell(row_it, col_idx, it_bit)
    if row_ib and col_idx:
        ws.cell(row_ib, col_idx, ib_bit)

    print(f"✅ 写入 {sheet_name} sheet, {gear} 列: it={it_bit}, ib={ib_bit}")

# 保存
wb.save("AIC8822_APC_6E_20250928V2_filled.xlsx")

