import pandas as pd
from openpyxl import load_workbook

# 读取表1
df = pd.read_excel("apc_raw_data.xlsx", sheet_name="all")

# 打开表2
wb = load_workbook("AIC8819H_APC_20250822_V1.xlsx")

for _, row in df.iterrows():
    sheet_name = row["ch_case"]
    gear = str(row["analog_gear"]).upper()   # 列名 (A, B, C ... 或数字转16进制)
    it_bit = row["wf_dac0_vl_it_bit"]
    ib_bit = row["wf_dac0_vl_ib_bit"]

    if sheet_name not in wb.sheetnames:
        print(f"⚠️ sheet {sheet_name} 不存在，跳过")
        continue

    ws = wb[sheet_name]

    # 找到行号（APC_BITS = wf_dtmx_dac_vl_it_bit / wf_dtmx_dac_vl_ib_bit）
    row_it = None
    row_ib = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val == "wf_dtmx_dac_vl_it_bit":
            row_it = r
        elif val == "wf_dtmx_dac_vl_ib_bit":
            row_ib = r

    # 找到列号（根据 analog_gear）
    col_idx = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value).upper() == gear:
            col_idx = c
            break

    if row_it and col_idx:
        ws.cell(row_it, col_idx, it_bit)
    if row_ib and col_idx:
        ws.cell(row_ib, col_idx, ib_bit)

    print(f"✅ 写入 {sheet_name} sheet, {gear} 列: it={it_bit}, ib={ib_bit}")

# 保存
wb.save("AIC8819H_APC_20250822_V1_filled.xlsx")

