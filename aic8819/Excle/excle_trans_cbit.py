import pandas as pd
from openpyxl import load_workbook

# 读取表1
df = pd.read_excel("./data/apc_raw_data_cbit.xlsx", sheet_name="all")

# 打开表2
wb = load_workbook("AIC8819H_APC_20250912.xlsx")

for _, row in df.iterrows():
    sheet_name = row["test_case"]
    gear = str(row["analog_gear"]).upper()   # 列名 (A, B, C ... 或数字转16进制)
    cbit = row["wf_dtmx_cbit"]


    if row["hb"] == 1:
        cbit_name = "wf_dtmx_cbit_hb"
    else:
        cbit_name = "wf_dtmx_cbit_lb"

    if sheet_name not in wb.sheetnames:
        print(f"⚠️ sheet {sheet_name} 不存在，跳过")
        continue

    ws = wb[sheet_name]

    # 找到行号（APC_BITS = wf_dtmx_dac_vl_cbit / wf_dtmx_dac_vl_ib_bit）
    row_cbit = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val == cbit_name:
            row_cbit = r

    # 找到列号（根据 analog_gear）
    col_idx = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value).upper() == gear:
            col_idx = c
            break

    if row_cbit and col_idx:
        ws.cell(row_cbit, col_idx, cbit)

    print(f"✅ 写入 {sheet_name} sheet, {gear} 列: it={cbit}")

# 保存
wb.save("AIC8819H_APC_20251021_filled.xlsx")

