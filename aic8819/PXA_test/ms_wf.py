import os
import sys
import openpyxl

from icbasic.aicinstr.rs.cmp180 import *
from icbasic.aicintf.uart import *
from MsPFM.csv import *
from MsPFM.GlobalVar import *

GX = GlobalVar()


class WF_MS_TABLE:
    def __init__(self, xlsx_path):
        self.xlsx_path = xlsx_path

    def read(self):
        ms_table = openpyxl.load_workbook(self.xlsx_path, data_only=True)
        l_test_lines = []
        if "PMS" in ms_table.sheetnames:
            pms_sheet = ms_table["PMS"]
            #print(pms_sheet.max_row)
            #print(pms_sheet.max_column)
            for rowx in range(2, pms_sheet.max_row+1):
                linex = []
                for columnx in range(1, pms_sheet.max_column+1):
                    linex.append(pms_sheet.cell(rowx, columnx).value)
                if None in linex[0:6]:
                    continue
                else:
                    l_test_lines.append(linex)
        else:
            print("ERROR: NO SHEET NAMED PMS FOUND!!!!")
        # print(l_test_lines)
        return l_test_lines


class WF_MS_LINE:
    def __init__(self, l_line):
        self.l_line = l_line
        self.index_offset = 4

    def boardno(self):
        return str(self.l_line[-4 + self.index_offset]).strip()

    def ant(self):
        return str(self.l_line[-3 + self.index_offset]).strip()

    def binversion(self):
        return str(self.l_line[-2 + self.index_offset]).strip()

    def route(self):
        return str(self.l_line[-1 + self.index_offset]).strip()

    def enable(self):
        return str(self.l_line[0 + self.index_offset]).strip()

    def l_ch(self):
        ch_string = str(self.l_line[1 + self.index_offset]).strip()
        if " " in ch_string:
            return list(range(int(ch_string.split(" ")[0]), int(ch_string.split(" ")[1]), int(ch_string.split(" ")[2])))
        else:
            return ch_string.split(",")

    def rate(self):
        return str(self.l_line[2 + self.index_offset]).strip()

    def bw(self):
        return str(self.l_line[3 + self.index_offset]).strip()

    def len(self):
        return str(self.l_line[4 + self.index_offset]).strip()

    def l_pwr(self):
        pwr_string = self.l_line[5 + self.index_offset].strip()
        if " " in pwr_string:
            return list(range(int(pwr_string.split(" ")[0]), int(pwr_string.split(" ")[2]), int(pwr_string.split(" ")[1])))
        elif ":" in pwr_string:
            return list(range(int(pwr_string.split(":")[0]), int(pwr_string.split(":")[2]), int(pwr_string.split(":")[1])))
        else:
            return pwr_string.split(",")

    def uart_cmd(self):
        return self.l_line[6 + self.index_offset]

    def res_pwr(self):
        return self.result_cell_check(self.l_line[7 + self.index_offset])

    def res_evm_avg(self):
        return self.result_cell_check(self.l_line[8 + self.index_offset])

    def res_evm_peak(self):
        return self.result_cell_check(self.l_line[9 + self.index_offset])

    def res_mask_avg(self):
        return self.result_cell_check(self.l_line[10 + self.index_offset])

    def res_mask_max(self):
        return self.result_cell_check(self.l_line[11 + self.index_offset])

    @staticmethod
    def result_cell_check(cell_value):
        if cell_value is None:
            return False
        elif cell_value.strip() in ["y", "Y", "yes", "YES", "Yes"]:
            return True
        else:
            return False

    def l_setch_ucmd(self):
        l_chx = []
        for chx in self.l_ch():
            l_chx.append("setch {}".format(chx))
        return l_chx

    def setrate_ucmd(self):
        return "setrate {}".format(self.rate())

    def setbw_ucmd(self):
        return "setbw {}".format(self.bw())

    def setlen_ucmd(self):
        return "setlen {}".format(self.len())

    def l_setpwr_ucmd(self):
        l_pwrx = []
        for pwrx in self.l_pwr():
            l_pwrx.append("setpwr {}".format(pwrx))
        return l_pwrx

    def l_uartcmd(self):
        if isinstance(self.uart_cmd(), str):
            if self.uart_cmd().strip() != "":
                return self.uart_cmd().split(",")
        else:
            return None


class WF_MS:
    def __init__(self, xlsx_path):
        self.db_table = WF_MS_TABLE(xlsx_path)
        self.l_test_lines = self.db_table.read()
        self.UARTX = GX.get_value("UARTX")
        self.CMPX = GX.get_value("CMPX")
        self.CSVX = GX.get_value("CSVX")

    def wf_ms_table(self):
        csv_header = "BoardNo, ANT, BinVersion, Channel, Rate, BandWidth, Length, SetPwr, CMD, MsPwrAvg, MsEvmAvg\n"
        self.CSVX.write_append_line(csv_header)
        for linex in self.l_test_lines:
            db_line = WF_MS_LINE(linex)
            if db_line.enable() not in ["Y", "y", "YES", "yes"]:
                continue
            # print(db_line.l_setch_ucmd())
            # print(db_line.l_setpwr_ucmd())
            # print(db_line.setbw_ucmd())
            # print(db_line.setrate_ucmd())
            self.UARTX.sendcmd("settx 1")
            self.UARTX.sendcmd("pwrmm 1")
            self.UARTX.sendcmd(db_line.setrate_ucmd())
            self.UARTX.sendcmd(db_line.setbw_ucmd())
            self.UARTX.sendcmd(db_line.setlen_ucmd())
            rate = " ".join(db_line.setrate_ucmd().strip().split(" ")[1:])
            bw = " ".join(db_line.setbw_ucmd().strip().split(" ")[1:])
            len = " ".join(db_line.setlen_ucmd().strip().split(" ")[1:])

            # set route
            if db_line.route() != "":
                self.CMPX.wlan_set_route(db_line.route())

            wlan_standard = "11ax"
            if rate.strip().split(" ")[0] == "5":
                wlan_standard = "11ax"
            elif rate.strip().split(" ")[0] == "4":
                wlan_standard = "11ac"
            elif rate.strip().split(" ")[0] == "2":
                wlan_standard = "11n"
            elif rate.strip().split(" ")[0] == "0":
                if rate.strip().split(" ")[1] == "3":
                    wlan_standard = "11b"
                else:
                    wlan_standard = "11g"

            self.CMPX.wlan_set_standard(wlan_standard)

            # print(bw)
            if "0 0" in bw:
                self.CMPX.wlan_set_bandwidth("20")
            elif "1 1" in bw:
                self.CMPX.wlan_set_bandwidth("40")
            elif "2 2" in bw:
                self.CMPX.wlan_set_bandwidth("80")

            for setchx in db_line.l_setch_ucmd():
                ch = setchx.strip().split(" ")[1]
                self.CMPX.wlan_set_freq_by_ch(ch)
                self.UARTX.sendcmd(setchx)
                time.sleep(2)
                for setpwrx in db_line.l_setpwr_ucmd():
                    setpwr = " ".join(setpwrx.strip().split(" ")[1:])
                    self.UARTX.sendcmd(setpwrx)

                    cmdx = ""
                    #print(db_line.l_uartcmd())
                    try:
                        if db_line.l_uartcmd() is not None:
                            for u_cmdx in db_line.l_uartcmd():
                                cmdx = cmdx + self.UARTX.sendcmd(u_cmdx)
                                #print(cmdx)
                    except:
                        cmdx = "ERROR"

                    for key_str in [",", "\n", "\r", "aic>"]:
                        cmdx = cmdx.replace(key_str, " ")

                    self.CMPX.wlan_auto_peak_pwr()
                    self.CMPX.wlan_meas_start()
                    time.sleep(2)

                    # measure pwr and evm
                    if wlan_standard == "11b":
                        if db_line.res_pwr():
                            ms_pwr = self.CMPX.wlan_meas_11b_pwr()
                        else:
                            ms_pwr = "NA"

                        if db_line.res_evm_avg():
                            ms_evm_avg = self.CMPX.wlan_meas_11b_evm_rms()
                        else:
                            ms_evm_avg = "NA"

                        if db_line.res_evm_peak():
                            ms_evm_peak = self.CMPX.wlan_meas_11b_evm_peak()
                        else:
                            ms_evm_peak = "NA"
                    else:
                        if db_line.res_pwr():
                            ms_pwr = self.CMPX.wlan_meas_pwr()
                        else:
                            ms_pwr = "NA"

                        if db_line.res_evm_avg():
                            ms_evm_avg = self.CMPX.wlan_meas_evm()
                        else:
                            ms_evm_avg = "NA"

                        if db_line.res_evm_peak():
                            ms_evm_peak = self.CMPX.wlan_meas_evm_peak()
                        else:
                            ms_evm_peak = "NA"

                    # measure mask
                    if db_line.res_mask_avg():
                        ms_mask_avg = self.CMPX.wlan_meas_tsmask_avg_maxval()
                    else:
                        ms_mask_avg = "NA"

                    if db_line.res_mask_max():
                        ms_mask_max = self.CMPX.wlan_meas_tsmask_max_maxval()
                    else:
                        ms_mask_max = "NA"

                    self.CMPX.wlan_meas_abort()

                    results = "{},{},{},{},{},{},{},{},{},{},{},{},{},{}".format(db_line.boardno(), db_line.ant(), db_line.binversion(), ch, rate,
                                                                     bw, len, setpwr, cmdx, ms_pwr, ms_evm_avg, ms_evm_peak, ms_mask_avg, ms_mask_max)
                    self.CSVX.write_append_line(results)
                    print(results)
        self.UARTX.sendcmd("settx 0")


if __name__ == "__main__":
    csv_path = "./MsDatas/AIC8822_WF_MEASURE_TABLE_20240627_V1.csv"
    xlsx_path = "./MsTables/WF_MEASURE_TABLE_20240627_V1.xlsx"

    CSVX = CSV(csv_path)
    csv_header = "Board NO, Bin, Channel, Rate, BandWidth, Length, SetPwr, CmdInfo, MsPwrAvg, MsEvmAvg, MsEvmPeak, " \
                 "MsMaskAvg, MsMaskPeak\n"
    CSVX.write_append_line(csv_header)

    UARTc = Uart(8)
    UARTc.open()

    host = "10.21.10.200"
    port = 5025
    CMPX = CMP180()
    CMPX.open_tcp(host, port)

    MS = WF_MS(xlsx_path)
    MS.wf_ms_table()

    # CMPX.wlan_set_bandwidth("80")