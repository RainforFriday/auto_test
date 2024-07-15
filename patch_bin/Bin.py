import os
import sys
import openpyxl


class Bin:
    def __init__(self, bin_path):
        self.path = bin_path

    def read_to_list(self):
        # read ref bin data
        ref_bin_file = open(self.path, "rb")
        ref_bin_size = os.path.getsize(self.path)
        ref_bin_data = []
        for i in range(ref_bin_size):
            ref_bin_data.append(ref_bin_file.read(1))
        ref_bin_file.close()
        return ref_bin_data

    def replace_list(self, l_old = [b'\x78', b'\x00', b'\x78', b'\x00'], l_new = [b'\xA0', b'\x00', b'\xA0', b'\x00'], max = 1):
        # replace cal ipa target current
        l_bin_data = self.read_to_list()
        flag_index = []
        for i in range(len(l_bin_data) - len(l_old)):
            if l_bin_data[i:i + len(l_old)] == l_old:
                flag_index.append(i)
            if len(flag_index) == max:
                break
        # print(flag_index)
        for index in flag_index:
            l_bin_data[index:index+len(l_old)] = l_new

        return l_bin_data

    def replace_apc(self, l_old_apc, l_new_apc):
        flag_index_list = self.search_table(l_old_apc)
        print(flag_index_list)
        l_new_bin = self.read_to_list()
        for flag_index in flag_index_list:
            l_new_bin[flag_index:flag_index+len(l_old_apc)] = l_new_apc
        return l_new_bin

    def replace_apcs(self, l_old_apcs, l_new_apcs):
        l_new_bin = self.read_to_list()
        for old_apc in l_old_apcs:
            x_index = l_new_apcs.index(old_apc)
            flag_index_list = self.search_table(old_apc)
            print(flag_index_list)
            for flag_index in flag_index_list:
                l_new_bin[flag_index:flag_index+len(old_apc)] = l_new_apcs[x_index]
        return l_new_bin


    def search_table(self, l_table):
        ref_bin_data = self.read_to_list()
        apc_table_bytes_list = l_table
        apc_flag_index_list = []
        for i in range(len(ref_bin_data) - len(apc_table_bytes_list)):
            if ref_bin_data[i:i + len(apc_table_bytes_list)] == apc_table_bytes_list:
                apc_flag_index_list.append(i)
        return apc_flag_index_list

    @staticmethod
    def write_bin(l_bin_data, bin_name):
        # write new bin to file
        new_bin_file = open(bin_name, "wb")
        for data_byte in l_bin_data:
            new_bin_file.write(data_byte)
        new_bin_file.close()


class APCXLSX:
    def __init__(self, apc_xlsx_path):
        self.path = apc_xlsx_path

    def read_table_to_list(self, sheet_name):
        apc_wb = openpyxl.load_workbook(self.path, data_only=True)
        apc_bytes_dict = {}
        if sheet_name in apc_wb.sheetnames:
            apc_sheet = apc_wb[sheet_name]
            value_bytes = []
            for rowx in range(2, 50):
                # for AIC8822, cilumn = 25
                # value_hex = apc_sheet.cell(row=rowx, column=25).value[2:-1]
                # for AIC8820, column = 24
                value_hex = apc_sheet.cell(row=rowx, column=24).value[2:-1]
                value_bytes = value_bytes + self.reg_value_to_bytes(value_hex)
            apc_bytes_dict[sheet_name] = value_bytes
        return apc_bytes_dict[sheet_name]

    @staticmethod
    def reg_value_to_bytes(reg_value):
        byte_list = []
        if len(reg_value) == 8:
            byte_list.append(bytes.fromhex(reg_value[6:8]))
            byte_list.append(bytes.fromhex(reg_value[4:6]))
            byte_list.append(bytes.fromhex(reg_value[2:4]))
            byte_list.append(bytes.fromhex(reg_value[0:2]))
        return byte_list


def bin_replace_apc(old_apc_xlsx_path, new_apc_xlxs_path, old_bin_path, new_bin_path, sheet_name):
    OLD_XLSX = APCXLSX(old_apc_xlsx_path)
    NEW_XLSX = APCXLSX(new_apc_xlxs_path)
    OLD_BIN = Bin(old_bin_path)

    L_OLD_APC = OLD_XLSX.read_table_to_list(sheet_name)
    L_NEW_APC = NEW_XLSX.read_table_to_list(sheet_name)

    l_new_bin_data = OLD_BIN.replace_apc(L_OLD_APC, L_NEW_APC)
    OLD_BIN.write_bin(l_new_bin_data, new_bin_path)


def bin_replace_apcs(old_apc_xlsx_path, new_apc_xlxs_path, old_bin_path, new_bin_path, sheet_names):
    OLD_XLSX = APCXLSX(old_apc_xlsx_path)
    NEW_XLSX = APCXLSX(new_apc_xlxs_path)
    OLD_BIN = Bin(old_bin_path)

    L_OLD_APCS = []
    L_NEW_APCS = []
    for sheet_name in sheet_names:
        L_OLD_APCS.append(OLD_XLSX.read_table_to_list(sheet_name))
        L_NEW_APCS.append(NEW_XLSX.read_table_to_list(sheet_name))

    l_new_bin_data = OLD_BIN.replace_apcs(L_OLD_APCS, L_NEW_APCS)
    OLD_BIN.write_bin(l_new_bin_data, new_bin_path)


if __name__ == "__main__":
    # 150: A0 00 A0 00
    # 150: 96 00 96 00
    # 140: 8C 00 8C 00
    # 130: 82 00 82 00
    # 120: 78 00 78 00
    # 110: 6E 00 6E 00
    # 100: 64 00 64 00
    # 90: 5A 00 5A 00
    # 80: 50 00 50 00
    #bin_path = "./testmode20_ipa_dbg.bin"
    #new_bin = "testmode20_ipa_160ma.bin"
    #AIC_BIN = Bin(bin_path)
    # print(AIC_BIN.read_to_list())
    #AIC_BIN.write_bin(new_bin)

    old_apc_xlsx = "./AIC8822_APC_LB_0618.xlsx"
    new_apc_xlsx = "./AIC8822_APC_LB_0620.xlsx"
    old_bin = "./testmode_8822_0619_2g4_1.bin"
    new_bin = "./testmode_8822_0620.bin"

    bin_replace_apc(old_apc_xlsx, new_apc_xlsx, old_bin, new_bin, "OFDM_H")
