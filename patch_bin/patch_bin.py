import os
import sys

import openpyxl


CHANNEL_LIST = ["CH42", "CH58", "CH106", "CH122", "CH138", "CH155"]
RATE_LIST = ["H", "M"]

ONE_APC_BYTES = 192
APC_NUM_LB = 2
APC_NUM_HB = 12


def apc_hb_sheet_names():
    apc_sheet_names_x = []
    for rate in RATE_LIST:
        for ch in CHANNEL_LIST:
            apc_sheet_names_x.append(ch+"_"+rate)
    return list(reversed(apc_sheet_names_x))


def apc_lb_sheet_names_lb():
    return ["OFDM_H", "11B"]


def get_apc_bytes_dict(apc_file_xlsx, apc_sheet_names):
    apc_wb = openpyxl.load_workbook(apc_file_xlsx, data_only=True)
    apc_bytes_dict = {}
    for apc_sheet_name in apc_sheet_names:
        if apc_sheet_name in apc_wb.sheetnames:
            apc_sheet = apc_wb[apc_sheet_name]
            value_bytes = []
            for rowx in range(2, 50):
                value_hex = apc_sheet.cell(row=rowx, column=25).value[2:-1]
                value_bytes = value_bytes + reg_value_to_bytes(value_hex)
            apc_bytes_dict[apc_sheet_name] = value_bytes
            """"
            print(len(value_bytes))
            for xx in value_bytes:
                print(xx.hex())
            """
    return apc_bytes_dict


def get_apc_bytes_list(apc_file_xlsx, apc_sheet_names):
    apc_data_dict = get_apc_bytes_dict(apc_file_xlsx, apc_sheet_names)
    apc_bytes_list = []
    for apc_sheet_name in apc_sheet_names:
        apc_bytes_list = apc_bytes_list + apc_data_dict[apc_sheet_name]
    return apc_bytes_list


def reg_value_to_bytes(reg_value):
    byte_list = []
    if len(reg_value) == 8:
        byte_list.append(bytes.fromhex(reg_value[6:8]))
        byte_list.append(bytes.fromhex(reg_value[4:6]))
        byte_list.append(bytes.fromhex(reg_value[2:4]))
        byte_list.append(bytes.fromhex(reg_value[0:2]))
    return byte_list


def patch_bin(ref_bin, apc_xlsx, new_bin, hb_en=True, apc_flag=b'\x0d\x0c\x0b\x0a\x24\x24\x55\x55'):

    # read ref bin data
    ref_bin_file = open(ref_bin, "rb")
    ref_bin_size = os.path.getsize(ref_bin)
    ref_bin_data = []
    for i in range(ref_bin_size):
        ref_bin_data.append(ref_bin_file.read(1))
    ref_bin_file.close()
    # print(ref_bin_data)

    # get apc_sheet_names
    if hb_en:
        apc_sheet_names = apc_hb_sheet_names()
    else:
        apc_sheet_names = apc_lb_sheet_names_lb()

    # read apc xlsx data
    apc_bytes_list = get_apc_bytes_list(apc_xlsx, apc_sheet_names)
    # print(len(apc_bytes_list))

    # search apc flag index
    # apc_flag_list = []
    # for i in range(len(apc_flag)):
    #     apc_flag_list.append(apc_flag[i])
    # print(apc_flag_list)
    apc_flag_list = [b'\x0d', b'\x0c', b'\x0b', b'\x0a', b'\x24', b'\x24', b'\x55', b'\x55']
    apc_flag_index = None
    for i in range(len(ref_bin_data) - len(apc_flag_list)):
        if ref_bin_data[i:i+len(apc_flag_list)] == apc_flag_list:
            apc_flag_index = i
            break
    # print 2 tablesi
    # print(apc_flag_index+8+192*2)
    # print(hex(apc_flag_index+8+192*2))
    # print(ref_bin_data[apc_flag_index+8:apc_flag_index+8+192*2])
    # print(ref_bin_data[apc_flag_index+8+192*2:apc_flag_index+8+192*3])

    # create new bin data
    new_bin_data = []
    if hb_en:
        table_num = APC_NUM_LB
    else:
        table_num = 0

    for i in range(len(ref_bin_data)):
        apc_update_start_index = apc_flag_index + len(apc_flag_list) + ONE_APC_BYTES*table_num
        # print(hex(apc_update_start_index))
        if i in range(apc_update_start_index, apc_update_start_index+len(apc_bytes_list)):
            new_bin_data.append(apc_bytes_list[i-apc_update_start_index])
        else:
            new_bin_data.append(ref_bin_data[i])

    # write new bin to file
    new_bin_file = open(new_bin, "wb")
    for data_byte in new_bin_data:
        new_bin_file.write(data_byte)
    new_bin_file.close()


def diff_bin(ref_bin, x_bin):
    ref_bin_file = open(ref_bin, "rb")
    ref_bin_size = os.path.getsize(ref_bin)
    ref_bin_data = []
    for i in range(ref_bin_size):
        ref_bin_data.append(ref_bin_file.read(1))
    ref_bin_file.close()

    x_bin_file = open(x_bin, "rb")
    x_bin_size = os.path.getsize(x_bin)
    x_bin_data = []
    for i in range(x_bin_size):
        x_bin_data.append(x_bin_file.read(1))
    x_bin_file.close()

    for i in range(len(ref_bin_data)):
        if ref_bin_data[i] != x_bin_data[i]:
            print("Index : " + hex(i) + ", ref bin: " + str(ref_bin_data[i]) + ", X bin: " + str(x_bin_data[i]))

    if len(ref_bin_data) != len(x_bin_data):
        print("Bin data size not the same!!!")
        return False


def search_table(ref_bin, apc_xlsx, hb_en=True, table_name="CH42_H"):
    ref_bin_file = open(ref_bin, "rb")
    ref_bin_size = os.path.getsize(ref_bin)
    ref_bin_data = []
    for i in range(ref_bin_size):
        ref_bin_data.append(ref_bin_file.read(1))
    ref_bin_file.close()

    # get apc_sheet_names
    if hb_en:
        apc_sheet_names = apc_hb_sheet_names()
    else:
        apc_sheet_names = apc_lb_sheet_names_lb()

    apc_table_bytes_list = get_apc_bytes_dict(apc_xlsx, apc_sheet_names)[table_name]

    apc_flag_index = None
    for i in range(len(ref_bin_data) - len(apc_table_bytes_list)):
        if ref_bin_data[i:i+len(apc_table_bytes_list)] == apc_table_bytes_list:
            apc_flag_index = i
            break
    print(apc_flag_index)


if __name__ == "__main__":
    ref_bin_path = "./testmode_8822_0618.bin"
    apc_xlsx_path = "./AIC8822_APC_LB_0618.xlsx"
    new_bin_path = "./testmode22_0618_lbVmbit1_10.bin"
    patch_bin(ref_bin_path, apc_xlsx_path, new_bin_path, hb_en=False)
    diff_bin(ref_bin_path, new_bin_path)
    """
    search_table(ref_bin_path, apc_xlsx_path, "CH42_H")
    search_table(ref_bin_path, apc_xlsx_path, "CH58_H")
    search_table(ref_bin_path, apc_xlsx_path, "CH106_H")
    search_table(ref_bin_path, apc_xlsx_path, "CH122_H")
    search_table(ref_bin_path, apc_xlsx_path, "CH138_H")
    search_table(ref_bin_path, apc_xlsx_path, "CH155_H")
    search_table(ref_bin_path, apc_xlsx_path, "CH42_M")
    search_table(ref_bin_path, apc_xlsx_path, "CH58_M")
    search_table(ref_bin_path, apc_xlsx_path, "CH106_M")
    search_table(ref_bin_path, apc_xlsx_path, "CH122_M")
    search_table(ref_bin_path, apc_xlsx_path, "CH138_M")
    search_table(ref_bin_path, apc_xlsx_path, "CH155_M")
    # print(xx)
    """

    """
    apc_flag = b'\x0d\x0c\x0b\x0a'
    bin_file_path = "./testmode_8822_0416_bin.bin"
    bin_file = open("./testmode_8822_0416_bin.bin", "rb")
    size = os.path.getsize("./testmode_8822_0416_bin.bin")
    for i in range(100):
        data = bin_file.read(1)
        print(data.hex())
    bin_file.close()
    with open(bin_file_path, "rb") as BIN:
        data = BIN.read()
        # print(data.hex())
        for line in BIN.readlines():
            print(line.hex())
            if apc_flag in line:
                print(line.hex())
                # print(line.replace(apc_flag, b'AICAICAIC'))
    """